#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Gera tabela-fretes.json a partir do BANCO_DE_DADOS.xlsx (base de cálculo da OBS).

Uso:
    python3 ferramentas/gerar-tabela.py CAMINHO/BANCO_DE_DADOS.xlsx

Estrutura de cada aba de transportadora:
  - Bloco preços (col A-E): Rota | Categoria de Veículo | Valor (R$) | Tipo | Dias
  - Bloco trajetos (col G-K): Rota | Origem Cidade | Origem UF | Destino Cidade | Destino UF
  Ligados pelo NOME da rota.
Aba "Configurações":
  - Categorias (col A)
  - Taxas por cidade (col C-F): Cidade | UF | Recebimento | Coleta/Entrega
  - (matriz RCTR-C já está embutida no index.html como SEGURO_TX)
"""
import sys, json, unicodedata
from openpyxl import load_workbook

def norm(s):
    if s is None: return ''
    s = str(s).replace('\xa0',' ').strip()
    s = ''.join(c for c in unicodedata.normalize('NFD', s) if unicodedata.category(c) != 'Mn')
    return s.lower().strip()

def numv(v):
    if v is None or v == '': return None
    if isinstance(v,(int,float)): return float(v)
    s = str(v).replace('\xa0','').replace('R$','').strip().replace('.','').replace(',','.')
    try: return float(s)
    except: return None

def main(path):
    wb = load_workbook(path, data_only=True)
    cfg = wb['Configurações']

    categorias = [cfg.cell(r,1).value for r in range(2,20) if cfg.cell(r,1).value]

    # Taxas por cidade (col C=3 cidade, D=4 UF, E=5 recebimento, F=6 coleta/entrega)
    cidades = {}
    for r in range(2, cfg.max_row+1):
        cid = cfg.cell(r,3).value
        if not cid: continue
        uf = (cfg.cell(r,4).value or '').strip().upper() if cfg.cell(r,4).value else ''
        rec = numv(cfg.cell(r,5).value)
        col = numv(cfg.cell(r,6).value)
        key = norm(cid)
        cidades[key] = { 'cidade': str(cid).strip(), 'uf': uf,
                         'recebimento': rec, 'coletaEntrega': col }

    rotas = []
    for ws in wb.worksheets:
        if ws.title == 'Configurações': continue
        transp = ws.cell(2,1).value or ws.title
        # preços por rota
        precos = {}   # rotaNome -> {'prazo':dias, 'valores':{cat:valor}}
        for r in range(7, ws.max_row+1):
            rota = ws.cell(r,1).value
            cat  = ws.cell(r,2).value
            val  = numv(ws.cell(r,3).value)
            prazo= ws.cell(r,5).value
            if not rota or not cat or val is None: continue
            rn = str(rota).strip()
            precos.setdefault(rn, {'prazo':None,'valores':{}})
            precos[rn]['valores'][str(cat).strip()] = val
            if prazo not in (None,'') and precos[rn]['prazo'] is None:
                try: precos[rn]['prazo'] = int(prazo)
                except: pass
        # trajetos por rota (col G=7 rota, H=8 oCidade, I=9 oUF, J=10 dCidade, K=11 dUF)
        trajetos = {}
        for r in range(7, ws.max_row+1):
            rota = ws.cell(r,7).value
            oc = ws.cell(r,8).value; ou = ws.cell(r,9).value
            dc = ws.cell(r,10).value; du = ws.cell(r,11).value
            if not rota or not oc or not dc: continue
            rn = str(rota).strip()
            trajetos.setdefault(rn, []).append({
                'o': norm(oc), 'oUF': (str(ou).strip().upper() if ou else ''),
                'd': norm(dc), 'dUF': (str(du).strip().upper() if du else ''),
                'oNome': str(oc).strip(), 'dNome': str(dc).strip()
            })
        for rn, pr in precos.items():
            rotas.append({
                'transportadora': str(transp).strip(),
                'rota': rn,
                'prazoDias': pr['prazo'],
                'valores': pr['valores'],
                'trajetos': trajetos.get(rn, [])
            })

    out = { 'categorias': categorias, 'cidades': cidades, 'rotas': rotas,
            'geradoDe': path.split('/')[-1] }
    dest = 'tabela-fretes.json'
    with open(dest,'w',encoding='utf-8') as f:
        json.dump(out, f, ensure_ascii=False, separators=(',',':'))
    nTraj = sum(len(r['trajetos']) for r in rotas)
    print(f'OK -> {dest}')
    print(f'  categorias: {len(categorias)} {categorias}')
    print(f'  cidades c/ taxa: {len(cidades)}')
    print(f'  rotas (transportadora x rota): {len(rotas)} | trajetos: {nTraj}')

if __name__ == '__main__':
    main(sys.argv[1] if len(sys.argv)>1 else 'BANCO_DE_DADOS.xlsx')
