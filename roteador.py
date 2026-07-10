#!/usr/bin/env python3
"""
Roteador de fretes sobre o banco de rotas (banco_dados.xlsx).

A malha é hub-and-spoke: cada aba (transportadora) tem uma tabela de preços
por rota/categoria e um mapa de cidades que pertencem a cada rota. Rotas que
não existem diretamente são montadas combinando trechos (as CARRETAS 1/2/3 do
orçamento) via caminho de menor custo.

Uso como módulo:
    from roteador import Roteador
    rt = Roteador("banco_dados.xlsx")
    resultado = rt.rotear("Manaus", "AM", "Fortaleza", "CE", "carro passeio")
"""
import re
import heapq
import unicodedata

import openpyxl

CATFALL = {
    "carro passeio": ["carro passeio", "carro pequeno", "carro grande"],
    "carro grande": ["carro grande", "carro passeio", "carro pequeno"],
    "moto ate 300cc": ["moto ate 300cc", "moto ate 700cc"],
    "moto ate 700cc": ["moto ate 700cc", "moto acima de 700cc", "moto ate 300cc"],
    "moto acima de 700cc": ["moto acima de 700cc", "moto ate 700cc"],
}


def norm(s):
    if s is None:
        return ""
    s = str(s).strip().lower()
    s = "".join(c for c in unicodedata.normalize("NFD", s)
                if unicodedata.category(c) != "Mn")
    return re.sub(r"\s+", " ", s).strip()


def _hub_node(endpoint):
    m = re.match(r"^(.*)\(\s*([A-Za-z]{2})\s*\)\s*$", endpoint.strip())
    if m:
        return (norm(m.group(1)), norm(m.group(2)))
    return (norm(endpoint), "")


class Roteador:
    def __init__(self, caminho="banco_dados.xlsx"):
        wb = openpyxl.load_workbook(caminho, data_only=True)
        transporters = [s for s in wb.sheetnames if s != "Configurações"]
        self.edges = {}         # (A,B) -> {cat_norm: (valor, dias, transp)}  (mais barato)
        self.city_to_hub = {}   # (cidade,uf) -> {hub_nodes}
        self.hubs = set()
        for s in transporters:
            ws = wb[s]
            transp = ws.cell(2, 1).value or s
            for r in range(7, ws.max_row + 1):
                rn = ws.cell(r, 1).value
                cat = ws.cell(r, 2).value
                val = ws.cell(r, 3).value
                dias = ws.cell(r, 5).value
                if rn and cat and val is not None and " - " in str(rn):
                    parts = str(rn).split(" - ")
                    A, B = _hub_node(parts[0]), _hub_node(parts[-1])
                    try:
                        v = float(val)
                    except (TypeError, ValueError):
                        continue
                    d = self.edges.setdefault((A, B), {})
                    ck = norm(cat)
                    if ck not in d or v < d[ck][0]:
                        d[ck] = (v, dias, transp)
            for r in range(7, ws.max_row + 1):
                rn = ws.cell(r, 7).value
                if not rn or " - " not in str(rn):
                    continue
                parts = str(rn).split(" - ")
                A, B = _hub_node(parts[0]), _hub_node(parts[-1])
                oc, ou = ws.cell(r, 8).value, ws.cell(r, 9).value
                dc, du = ws.cell(r, 10).value, ws.cell(r, 11).value
                if oc:
                    self.city_to_hub.setdefault((norm(oc), norm(ou)), set()).update([A, B])
                if dc:
                    self.city_to_hub.setdefault((norm(dc), norm(du)), set()).update([A, B])
        for (A, B) in self.edges:
            self.hubs.add(A)
            self.hubs.add(B)

    def resolver_cidade(self, cidade, uf=""):
        cn, un = norm(cidade), norm(uf)
        cand = [h for h in self.hubs if h[0] == cn and (not un or h[1] == un)]
        if cand:
            return cand, "hub-exato"
        cand = [h for h in self.hubs
                if h[0] and (cn in h[0] or h[0] in cn) and (not un or h[1] == un)]
        if cand:
            return cand, "hub-aprox"
        for (c, u), hs in self.city_to_hub.items():
            if c == cn and (not un or u == un):
                return list(hs), "trajeto"
        for (c, u), hs in self.city_to_hub.items():
            if c and (cn in c or c in cn) and (not un or u == un):
                return list(hs), "trajeto-aprox"
        return [], "nao-encontrado"

    def _preco_trecho(self, A, B, cat_norm):
        d = self.edges.get((A, B))
        if not d:
            return None
        for c in CATFALL.get(cat_norm, [cat_norm]):
            if c in d:
                return d[c] + (c,)
        return None

    def rotear(self, origem, origem_uf, destino, destino_uf, categoria):
        cat = norm(categoria)
        src, sw = self.resolver_cidade(origem, origem_uf)
        dst, dw = self.resolver_cidade(destino, destino_uf)
        out = {"src_how": sw, "dst_how": dw, "legs": None, "total": None, "prazo": None}
        if not src or not dst:
            return out
        dst_set = set(dst)
        pq = [(0, [n], []) for n in src]
        heapq.heapify(pq)
        seen = {}
        while pq:
            cost, path, legs = heapq.heappop(pq)
            node = path[-1]
            if node in dst_set:
                out["total"] = cost
                out["legs"] = [
                    {"de": f"{A[0]}/{A[1]}", "para": f"{B[0]}/{B[1]}",
                     "valor": v, "dias": dias, "transp": tp, "cat": uc}
                    for (A, B, v, dias, tp, uc) in legs]
                dias_conhec = [l["dias"] for l in out["legs"] if l["dias"]]
                out["prazo"] = sum(d for d in dias_conhec if d)
                out["prazo_parcial"] = any(l["dias"] is None for l in out["legs"])
                return out
            if seen.get(node, 1e18) <= cost:
                continue
            seen[node] = cost
            if len(path) > 4:
                continue
            for (A, B) in self.edges:
                if A != node:
                    continue
                pl = self._preco_trecho(A, B, cat)
                if pl is None:
                    continue
                v, dias, tp, uc = pl
                heapq.heappush(pq, (cost + v, path + [B], legs + [(A, B, v, dias, tp, uc)]))
        return out


# Heurística simples de categoria a partir do modelo do veículo
_MOTO_KW = ["moto", "daytona", "sahara", "xre", "cb", "cg", "fan", "biz",
            "titan", "fazer", "hornet", "bros", "pop", "factor"]
_GRANDE_KW = ["van", "sprinter", "master", "ducato", "chassi", "caminh",
              "hr", "bongo", "kombi", "utilit"]


def categoria_do_modelo(modelo):
    m = norm(modelo)
    if any(k == m or (" " + k + " ") in (" " + m + " ") for k in _MOTO_KW):
        cc = re.search(r"(\d{3,4})", m)
        n = int(cc.group(1)) if cc else 0
        if n and n <= 300:
            return "moto ate 300cc"
        if n and n <= 700:
            return "moto ate 700cc"
        if n > 700:
            return "moto acima de 700cc"
        return "moto ate 700cc"
    if any(k in m for k in _GRANDE_KW):
        return "carro grande"
    return "carro passeio"
