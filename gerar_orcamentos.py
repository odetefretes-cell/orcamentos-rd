#!/usr/bin/env python3
"""
Gera um arquivo de orçamento (.xlsx) por lead do RD Station CRM, já com o
frete calculado a partir do banco de rotas (banco_dados.xlsx).

Fluxo:
  1. Puxa negociações via API do RD Station CRM.
  2. Filtra os leads no estágio "EMITIR ORÇAMENTO".
  3. Para cada lead com dados, preenche uma cópia do template
     (aba "ORÇAMENTO PARTICULAR" ou "CNPJ" conforme Empresa/particular):
       - cliente, telefone, origem, destino, veículo, observações;
       - CARRETAS (custo de cada perna) e prazo, roteando pela malha do banco.
  4. Deixa em branco (para fechamento manual): margem/LUCRO (N22), FIPE (F17),
     coleta/entrega (K10/K11) e CNPJ (não existe no CRM).

Uso:
    RD_TOKEN=seu_token python3 gerar_orcamentos.py

Requer: openpyxl  (pip install openpyxl)  e  roteador.py + banco_dados.xlsx
"""
import os
import re
import json
import urllib.request

from roteador import Roteador, categoria_do_modelo, norm

TOKEN = os.environ.get("RD_TOKEN", "")
TEMPLATE = os.environ.get("TEMPLATE", "emissor_orcamento.xlsx")
BANCO = os.environ.get("BANCO", "banco_dados.xlsx")
OUTDIR = os.environ.get("OUTDIR", "orcamentos_gerados")
STAGE_ALVO = "EMITIR ORÇAMENTO"
MAX_PAGES = int(os.environ.get("MAX_PAGES", "4"))

# Ajustes manuais de categoria por modelo (quando a heurística não basta)
CATEGORIA_MANUAL = {
    "sahara 300": "moto ate 300cc",
    "daytona 660": "moto ate 700cc",
}

# UF a partir de texto livre da cidade
UF = {"sp": "sp", "mg": "mg", "rs": "rs", "pe": "pe", "ba": "ba", "ac": "ac",
      "am": "am", "ce": "ce", "rj": "rj", "ro": "ro", "go": "go", "pa": "pa",
      "pr": "pr", "mt": "mt", "ms": "ms", "rn": "rn", "se": "se", "pb": "pb",
      "es": "es", "to": "to", "al": "al", "df": "df", "ap": "ap",
      "sao paulo": "sp", "minas gerais": "mg", "rio grande do sul": "rs",
      "pernambuco": "pe", "bahia": "ba", "acre": "ac", "amazonas": "am",
      "ceara": "ce", "rio de janeiro": "rj"}


def fetch_deals():
    deals = []
    for page in range(1, MAX_PAGES + 1):
        url = (f"https://crm.rdstation.com/api/v1/deals"
               f"?token={TOKEN}&limit=50&page={page}")
        with urllib.request.urlopen(url) as r:
            data = json.load(r)
        deals += data["deals"]
        if not data.get("has_more"):
            break
    return deals


def cf(deal, label):
    for c in deal.get("deal_custom_fields", []):
        if c["custom_field"]["label"] == label:
            return c["value"]
    return None


def sanit(s):
    return (re.sub(r"[^A-Za-z0-9]+", "_", s).strip("_"))[:40] or "lead"


def parse_cidade(raw):
    """Extrai (cidade, uf) de um texto livre."""
    n = norm(raw)
    n = re.sub(r"[.,\d]", " ", n)
    n = re.sub(r"\s+", " ", n).strip(" -")
    uf = ""
    for w in sorted(UF, key=lambda x: -len(x)):
        if n.endswith(" " + w) or n == w:
            uf = UF[w]
            n = n[:len(n) - len(w)].strip(" -")
            break
    return n.strip(" -"), uf


def gerar():
    import openpyxl

    if not TOKEN:
        raise SystemExit("Defina a variável de ambiente RD_TOKEN.")
    os.makedirs(OUTDIR, exist_ok=True)
    rt = Roteador(BANCO)

    deals = fetch_deals()
    alvo = [d for d in deals if d["deal_stage"]["name"].strip() == STAGE_ALVO]

    gerados, pulados = [], []
    for x in alvo:
        origem = cf(x, "Cidade de origem")
        destino = cf(x, "Cidade de destino")
        modelo = cf(x, "Modelo do veículo")
        if not any([origem, destino, modelo]):
            pulados.append(x["name"].replace(" - Negociação", "").strip())
            continue

        name = x["name"].replace(" - Negociação", "").strip()
        tipo = (cf(x, "Empresa/particular:") or "").strip().lower()
        sheet_name = "CNPJ" if tipo == "empresa" else "ORÇAMENTO PARTICULAR"

        wb = openpyxl.load_workbook(TEMPLATE)
        ws = wb[sheet_name]
        for sn in list(wb.sheetnames):
            if sn != sheet_name:
                del wb[sn]
        wb.active = 0

        # Cliente
        contato = (x.get("contacts") or [{}])[0]
        tel = ""
        if contato.get("phones"):
            tel = (contato["phones"][0].get("phone") or "").strip()
        ws["A10"] = f"Cliente: {name}" + (f"   |   Tel: {tel}" if tel else "")
        if sheet_name == "CNPJ":
            ws["A11"] = "CNPJ: "

        # Veículo
        ws["B17"] = modelo or ""
        ws["F17"] = None            # FIPE em branco
        func = (cf(x, "Veículo está funcional?") or "").strip().lower()
        obs = "FUNCIONA" if func == "sim" else (
            "NÃO FUNCIONA" if func in ("não", "nao") else "")
        extra = []
        envio = cf(x, "Data de envio do veículo")
        if envio:
            extra.append(f"Envio: {envio}")
        valor = cf(x, "Valor do Veículo")
        if valor:
            extra.append(f"Valor informado: {valor}")
        ws["H17"] = (obs + ("  |  " if obs and extra else "")
                     + "  ".join(extra)).strip()

        # Rota
        ws["B23"] = origem or ""
        ws["E23"] = destino or ""

        # Limpa frete/margem/prazo
        for c in ["K7", "K8", "K9", "N22", "K10", "K11"]:
            ws[c] = None
        ws["H23"] = None

        # Roteamento (CARRETAS = custo de cada perna; margem fica em branco)
        oc, ouf = parse_cidade(origem or "")
        dc, duf = parse_cidade(destino or "")
        cat = CATEGORIA_MANUAL.get(norm(modelo)) or categoria_do_modelo(modelo or "")
        r = rt.rotear(oc, ouf, dc, duf, cat)
        status = "SEM ROTA (frete em branco)"
        if r.get("legs"):
            for i, leg in enumerate(r["legs"][:3]):
                ws[["K7", "K8", "K9"][i]] = leg["valor"]
            prazo = r.get("prazo") or 0
            parcial = r.get("prazo_parcial")
            ws["H23"] = (f"{prazo} dias" if prazo else "a confirmar") + \
                        (" (confirmar)" if parcial else "")
            rota = " + ".join(f"{l['de'].split('/')[0]}→{l['para'].split('/')[0]}"
                              for l in r["legs"])
            status = f"R$ {r['total']:.0f} ({len(r['legs'])}x) | {rota}"

        fn = os.path.join(OUTDIR, f"Orcamento_{sanit(name)}.xlsx")
        wb.save(fn)
        gerados.append((name, cat, status))

    print(f"Gerados: {len(gerados)}")
    for name, cat, st in gerados:
        print(f"  • {name} [{cat}] -> {st}")
    if pulados:
        print("\nPulados (sem dados de rota/veículo):", pulados)


if __name__ == "__main__":
    gerar()
